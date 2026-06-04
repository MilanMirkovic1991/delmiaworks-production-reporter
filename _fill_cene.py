"""Fill STD_COST in Cene.xlsx with best-effort EUR prices per stocking UoM.

Strategy:
- Fill IN-class commercial items (hardware, sealants, paints, polymers, sheet/strip, wires) + 2 PK items.
- Leave blank: PP (internal semi-finished), FP (military finished),
  IN items that are military explosives/propellants (no public price),
  internal igniter/fuze components, and items with insufficient description.
- Units: EUR per stocking UoM (piece, kg, L, m, m^2 depending on item).
"""
from openpyxl import load_workbook

SRC = "Cene.xlsx"
DST = "Cene_popunjen.xlsx"

# Row -> (price_eur, unit_note). Rows omitted = leave STD_COST blank.
prices = {
    2:  (18.00, "EUR/kg, 2K PU paint base"),
    3:  (22.00, "EUR/kg, 2K PU paint hardener"),
    7:  (0.06,  "EUR/pc, M4x25 DIN"),
    9:  (0.04,  "EUR/pc, M5x6"),
    10: (0.02,  "EUR/pc, washer"),
    13: (0.08,  "EUR/pc, cable lug DIN 46234A"),
    14: (22.00, "EUR/kg, Sn60 leaded solder wire"),
    15: (8.00,  "EUR/kg, black marking paint"),
    16: (3.50,  "EUR/L, thinner"),
    17: (13.00, "EUR/cartridge, Teroson MS 931 ~290-310ml"),
    18: (35.00, "EUR/cartridge, Loctite 5366"),
    19: (6.00,  "EUR/kg, bitumen paint"),
    20: (3.00,  "EUR/L, bitumen thinner"),
    21: (0.30,  "EUR/m, ETF-K 16x0.2 wire"),
    22: (0.25,  "EUR/m, ETF-K 16x0.125 wire"),
    26: (2.27,  "EUR/kg, Bayblend FR3010 Covestro (plastic-price.com 2025)"),
    27: (2.50,  "EUR/kg, seamless steel pipe 139.7x22"),
    28: (18.00, "EUR/pc, wooden pallet G2000"),
    29: (35.00, "EUR/pc, bound wooden crate (estimate)"),
    31: (3.50,  "EUR/kg, sealing wire"),
    32: (0.03,  "EUR/pc, lead/Al seal"),
    36: (0.05,  "EUR/pc, M5x5 DIN 915 set screw"),
    37: (0.05,  "EUR/pc, M5x12"),
    38: (0.05,  "EUR/pc, plastic protective cap"),
    39: (0.15,  "EUR/pc, O-ring"),
    41: (0.30,  "EUR/m, ETF-K 035 16x0.20 wire"),
    43: (0.40,  "EUR/m, 3mm insulating sleeve"),
    44: (2.00,  "EUR/L, ethyl alcohol"),
    46: (12.00, "EUR/kg, shellac"),
    50: (6.00,  "EUR/kg, specialty paper Hammermill 0.3"),
    52: (1.10,  "EUR/kg, DC04 strip"),
    53: (8.00,  "EUR/m2, 1mm cork sheet"),
    54: (2.00,  "EUR/L, ethyl alcohol SRPS"),
    55: (27.00, "EUR/pc, Loctite 243 50ml (euro-industry.com)"),
    56: (14.00, "EUR/cartridge, Terostat MS 9220 310ml"),
    60: (1.10,  "EUR/kg, DC04 3mm sheet"),
    64: (3.50,  "EUR/kg, Al99.5 H111 sheet"),
    66: (3.50,  "EUR/L, thinner KOV-KZ-165"),
    67: (20.00, "EUR/kg, silicone thermal-insulating paint"),
    68: (2.80,  "EUR/kg, AlMg3 EN AW-5754 sheet"),
    73: (0.02,  "EUR/pc, 4mm chrome steel ball DIN 5401"),
    83: (0.03,  "EUR/pc, M2x8"),
    91: (8.00,  "EUR/kg, preservation grease ZP-3"),
    94: (0.025, "EUR/pc, 3mm stainless steel ball DIN 5401"),
    97: (12.00, "EUR/cartridge, Loctite neutral silicone sealant"),
    98: (1.40,  "EUR/kg, 11SMn30 free-cutting drawn bar"),
    112:(30.00, "EUR/m2, calf box leather"),
    114:(5.00,  "EUR/kg, foam rubber"),
    115:(20.00, "EUR/kg, 0.2mm celluloid foil"),
    116:(8.00,  "EUR/kg, CuZn30 brass strip"),
    119:(1.10,  "EUR/kg, DC03 strip"),
}

wb = load_workbook(SRC)
ws = wb["Export Worksheet"]

filled = 0
for row, (price, note) in prices.items():
    cell = ws.cell(row=row, column=5)  # STD_COST is column E
    cell.value = float(price)
    cell.number_format = "#,##0.00"
    filled += 1

wb.save(DST)
print(f"Saved {DST}: filled {filled} rows out of {ws.max_row - 1} data rows.")
print(f"Blank rows (military/internal): {ws.max_row - 1 - filled}")
